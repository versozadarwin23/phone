try:
    from urllib.request import urlretrieve
except ImportError:
    from urllib import urlretrieve

while True:
    try:
        urlretrieve('https://raw.githubusercontent.com/versozadarwin23/phone/main/phone.py', 'phone.py')
        break
    except:
        pass

from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.common.actions.pointer_input import PointerInput
from selenium.webdriver.common.actions import interaction
import time
import subprocess
import random
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from concurrent.futures import ThreadPoolExecutor

input_file = "devices.xlsx"
sheet_names = ["Phones", "Apps", "Comments", "Timeline", "Friends", "Reactions"]

# Load the relevant sheets from the Excel file
wb = load_workbook(input_file)

def get_sheet_data(sheet_name):
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))  # Skip header row
    headers = [cell.value for cell in sheet[1]]  # Extract headers
    return [dict(zip(headers, row)) for row in rows]

df_phones = get_sheet_data(sheet_names[0])
df_apps = get_sheet_data(sheet_names[1])
df_comments = get_sheet_data(sheet_names[2])
df_timeline = get_sheet_data(sheet_names[3])
df_friends = get_sheet_data(sheet_names[4])
df_reactions = get_sheet_data(sheet_names[5])

device_ids = [row['deviceID'] for row in df_phones]
android_device_serials = [row['androidDeviceSerial'] for row in df_phones]

def fetch_random_comment_by_category(category, number):
    comment_list = [row["Comment"] for row in df_comments if row["Category"] == category]
    return " ".join(random.sample(comment_list, k=number)) if len(comment_list) >= number else ""

def fetch_app_name_by_device_id(device_id):
    return [row for row in df_apps if row["deviceID"] == device_id]

def get_reaction_by_link(link):
    for row in df_reactions:
        if row["Link"] == link:
            return row["Reaction"]
    return ""

def adb_command(serial, command):
    try:
        return subprocess.check_output(f"adb -s {serial} {command}", shell=True)
    except subprocess.CalledProcessError as e:
        print(f"ADB Error on device {serial}: {e}")
        return None

def airplane_mode_off(device):
    if device.get("androidDeviceSerial"):
        adb_command(device['androidDeviceSerial'], "shell settings put global airplane_mode_on 0")
        adb_command(device['androidDeviceSerial'], "shell am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false")

def airplane_mode_on(device):
    if device.get("androidDeviceSerial"):
        adb_command(device['androidDeviceSerial'], "shell settings put global airplane_mode_on 1")
        adb_command(device['androidDeviceSerial'], "shell am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true")

def handle_device_tasks(device_id, android_device_serial):
    global dawssw_text, dawdsadwadasdwadwaferfgregre
    apps = fetch_app_name_by_device_id(device_id=device_id)
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-notifications")
    options.add_experimental_option('androidPackage', 'com.android.chrome')
    options.add_experimental_option('androidDeviceSerial', android_device_serial)

    while True:
        try:
            driver = webdriver.Chrome(executable_path="1.exe", options=options)
            break
        except:
            handle_device_tasks(device_id, android_device_serial)
    for x in apps:
        # show keyboard
        subprocess.check_output(
            "adb -s " + " " + android_device_serial + " " + "shell ime set com.emoji.keyboard.touchpal/com.cootek.smartinput5.TouchPalIME",
            shell=True)

        # airplane mode on
        subprocess.check_output(
            "adb -s " + " " + android_device_serial + " " + "shell settings put global airplane_mode_on 1", shell=True)
        subprocess.check_output(
            "adb -s " + " " + android_device_serial + " " + "shell am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true",
            shell=True)
        time.sleep(5)

        # airplane mode off
        subprocess.check_output(
            "adb -s " + " " + android_device_serial + " " + "shell settings put global airplane_mode_on 0", shell=True)
        subprocess.check_output(
            "adb -s " + " " + android_device_serial + " " + "shell am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false",
            shell=True)
        time.sleep(5)
        while True:
            try:
                driver.get("https://m.facebook.com/")
                break
            except:
                handle_device_tasks(device_id, android_device_serial)

        try:

            WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.NAME, "email")))
        except:
            handle_device_tasks(device_id, android_device_serial)

        try:
            WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.NAME, "email"))).send_keys(x.get("username"))
        except:
            handle_device_tasks(device_id, android_device_serial)

        while True:
            try:
                WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.NAME, "pass"))).send_keys(x.get("password"))
                break
            except:
                handle_device_tasks(device_id, android_device_serial)
        while True:
            try:
                WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, '[aria-label="Log in"]'))).click()
                break
            except:
                handle_device_tasks(device_id, android_device_serial)

        try:
            WebDriverWait(driver, 15).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, '[aria-label="Save"]'))).click()
            time.sleep(15)
        except:
            print(x["deviceID"] + " " + x.get(["username"]) + " " + "Login Error")

        try:
            WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, '[aria-label="Make a Post on Facebook"]')))
        except:
            print(x["deviceID"] + " " + x.get(["username"]) + " " + "Login Error")
            continue


        if 'reaction' in x and x["reaction"] == "yes":
            for link in x.get("links_to_react", "").split(" "):
                reaction = get_reaction_by_link(link=link)
                actions = ActionChains(driver)
                driver.get(link)
                try:
                    WebDriverWait(driver, 10).until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, '[aria-label="Like, button double tap and hold for more reactions."]')))
                except:
                    try:
                        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    except:
                        pass

                try:
                    dawdsadwadasdwadwaferfgregre = WebDriverWait(driver, 10).until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, '[aria-label="Like, button double tap and hold for more reactions."]')))[0]
                except:
                    try:
                        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    except:
                        pass

                actions.move_to_element(dawdsadwadasdwadwaferfgregre).perform()

                try:
                    dawssw = WebDriverWait(driver, 10).until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, 'div[style="color:#4b4c4f;"]')))[0]
                    dawssw_text = dawssw.text
                except:
                    pass

                try:
                    holdss = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div[aria-label='" + dawssw_text + "like, double tap and hold for more reactions'")))
                except:
                    continue
                touch_input = PointerInput(interaction.POINTER_TOUCH, 'touch')
                actions.w3c_actions = ActionBuilder(driver, mouse=touch_input)
                actions.w3c_actions.pointer_action.click_and_hold(holdss)
                actions.perform()
                time.sleep(3)
                actions.reset_actions()
                try:
                    WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div[aria-label='" + reaction + "']"))).click()
                    print(x["deviceID"] + " " + x["profile"] + " " + link + " " + reaction + " " + "React Done")
                    time.sleep(5)
                except:
                    pass

        if 'Comment' in x and x["Comment"] == "yes":
            actions = ActionChains(driver)
            for comments in x.get("comment_link", "").split(" "):
                send_comments = fetch_random_comment_by_category(category=x["category"], number=1)
                while True:
                    try:
                        time.sleep(5)
                        comms = WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'div[data-mcomponent="MInputBox"]')))
                        time.sleep(0.3)
                        actions.send_keys_to_element(comms, send_comments).perform()
                        time.sleep(0.3)
                        break
                    except:
                        handle_device_tasks(device_id, android_device_serial)

                while True:
                    try:
                        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, '[aria-label="Post a comment"]'))).click()
                        break
                    except:
                        try:
                            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, '[value="Comment"]'))).click()
                            break
                        except:
                            pass
                while True:
                    try:
                        WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, '[style="color:#65676b;"]')))
                        time.sleep(5)
                        print(x["deviceID"] + " " + x["profile"] + " " + comments + " " + "Comment done")
                        break
                    except:
                        pass

        if 'share' in x and x["share"] == "yes":
            actionsdawdaw = ActionChains(driver)
            for share in x.get("links_to_share", "").split(" "):
                subprocess.check_output("adb -s " + " " + android_device_serial + " " + "shell ime set com.emoji.keyboard.touchpal/com.cootek.smartinput5.TouchPalIME",shell=True)
                driver.get('https://m.facebook.com/composer/')
                whats_on_your_mind_click = WebDriverWait(driver, 10).until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, 'div[style="color:#666666;"]')))
                whats_on_your_mind = whats_on_your_mind_click[0]
                time.sleep(3)
                actionsdawdaw.send_keys_to_element(whats_on_your_mind, share).perform()
                time.sleep(3)
                try:
                    subprocess.check_output("adb -s" + " " + android_device_serial + " " + "shell input keyevent 4",shell=True)
                except:
                    pass
                subprocess.check_output("adb -s " + " " + android_device_serial + " " + "shell ime set com.wparam.nullkeyboard/.NullKeyboard",shell=True)
                time.sleep(5)
                whats_on_your_mind_clicks = WebDriverWait(driver, 10).until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, 'div[style="color:#666666;"]')))
                whats_on_your_mind_clear = whats_on_your_mind_clicks[0]
                actiondawdaws = ActionChains(driver)
                actiondawdaws.click(whats_on_your_mind_clear).perform()
                while True:
                    try:
                        time.sleep(0.3)
                        subprocess.check_output("adb -s " + " " + android_device_serial + " " + "shell input keyevent --longpress 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67 67", shell=True)
                        break
                    except:
                        pass

                time.sleep(0.3)
                postss_click = WebDriverWait(driver, 10).until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, 'div[style="color:#1877f2;"]')))
                time.sleep(0.3)
                actionsdawdaw.click(postss_click[0]).perform()
                print(x["deviceID"] + " " + x["profile"] + " " + share + " " + "Share done")

        while True:
                try:
                    driver.delete_all_cookies()
                    break
                except:
                    handle_device_tasks(device_id, android_device_serial)

def run_on_multiple_devices():
    with ThreadPoolExecutor(max_workers=len(device_ids)) as executor:
        executor.map(lambda device_id: handle_device_tasks(device_id, android_device_serials[device_ids.index(device_id)]), device_ids)

# Start task execution
run_on_multiple_devices()
