from __future__ import division
import itertools
import os

from xlrd import open_workbook
from openpyxl import load_workbook
from future.utils import iteritems
from past.builtins import xrange
from datetime import datetime
from multiprocessing import Process
import itertools
import sys


import random
input_file = "profile_settings.xlsx"
sheet_names=["Settings", "Links"]
sheet=""
all_data =[]


class XlToDict():
    def convert_sheet_to_dict(self, file_path=input_file, sheet=sheet, filter_variables_dict=None):
        """
        This method will convert excel sheets to dict. The input is path to the excel file or a sheet object.
        if file_path is None, sheet object must be provded. This method will convert only the first sheet.
        If you need to convert multiple sheets, please use the method fetch_data_by_column_by_sheet_name_multiple() and
        fetch_data_by_column_by_index_multiple().If you need to filter data by a specific keyword, specify the dict in
        filter_variables_dict like {column name : keyword} . Any rows that matches the keyword in the specified column
        will be returned. Multiple keywords can be specified.

        :param file_path: The path to the spreadsheet
        :param sheet: The sheet object from the spreadsheet
        :param filter_variables_dict: Optional. The keyword dict to filter out the data. specified as
        {column name : keyword}
        :return result_dict_list: List of dictionaries .
        """
        # read headers
        sheet_index = 0
        if file_path is not None:
            workbook = open_workbook(filename=file_path)
        keys = [sheet.cell(0, col_index).value for col_index in xrange(sheet.ncols)]

        found_row_dict_list = []
        for column_index, key in enumerate(keys):
            if filter_variables_dict is not None:
                for column_name, column_value in iteritems(filter_variables_dict):
                    if key == column_name:
                        for row_index in xrange(sheet.nrows):
                            if not (column_name == None and column_value == None):
                                if (sheet.cell(row_index, column_index).value) == column_value:
                                    found_row_dict = {
                                        keys[col_index_internal]: sheet.cell(row_index, col_index_internal).value
                                        for col_index_internal in xrange(sheet.ncols)}
                                    found_row_dict_list.append(found_row_dict)
                            else:
                                found_row_dict = {
                                    keys[col_index_internal]: sheet.cell(row_index, col_index_internal).value
                                    for col_index_internal in xrange(sheet.ncols)}
                                found_row_dict_list.append(found_row_dict)
            elif filter_variables_dict == {} or filter_variables_dict is None:
                filter_variables_dict = {}
                for row_index in xrange(sheet.nrows):
                    found_row_dict = {keys[col_index_internal]: sheet.cell(row_index, col_index_internal).value
                                      for col_index_internal in xrange(sheet.ncols)}
                    found_row_dict_list.append(found_row_dict)
                del found_row_dict_list[0]
        result_dict_list = []
        if len(found_row_dict_list) > 1 and len(filter_variables_dict) > 1:
            for a, b in itertools.combinations(found_row_dict_list, len(filter_variables_dict)):
                if a == b:
                    result_dict_list.append(a)
        else:
            result_dict_list = found_row_dict_list

        return result_dict_list

    def fetch_data_by_column_by_sheet_name_multiple(self, file_name, filter_variables_dict=None, sheet_names=sheet_names):
        """
        This method will convert multiple sheets in the excel file to dict. The input is path to the excel file .
        If sheet_names is not provided, this method will convert ALL the sheets.If you need to filter data by a specific
        keyword / keywords, specify the dict in filter_variables_dict like {column name : keyword} .
        Any rows that matches the keyword  in the specified column will be returned. Multiple keywords can be specified.

        :param file_path: The path to the spreadsheet
        :param sheet_names: Optional. The list of sheet names in the spreadsheet to convert. If not specified, will
        convert all the sheets
        :param filter_variables_dict: Optional. The keyword dict to filter out the data. specified as
        {column name-1 : keyword1, column name-2 : keyword2}
        :return result_dict_list: List of dictionaries .
        """
        workbook = open_workbook(filename=file_name)
        resultdictlist = []
        if sheet_names is None:
            sheet_names = workbook.sheet_names()

        for sheet_name in sheet_names:
            sheet = workbook.sheet_by_name(sheet_name)

            all_data.append(self.convert_sheet_to_dict(sheet=sheet, filter_variables_dict=filter_variables_dict))


        return all_data


myxlobject= XlToDict()
c = myxlobject.fetch_data_by_column_by_sheet_name_multiple(file_name=input_file,
                                                       sheet_names=sheet_names,
                                                       filter_variables_dict=None)


settings_sheets =  all_data[0]
links_sheet = all_data[1]


# print links_sheet


number_of_accounts = range(1,int(settings_sheets[0]["Accounts"]+1))




final_list_react = []
list_accts = []
for i in links_sheet:
    accounts_that_will_react = random.sample(number_of_accounts,int(i["Reactions"]))
    for x in accounts_that_will_react:
        d = {}
        d["account"] = x
        d["link"] = i["Links"]
        list_accts.append(d)
    final_list_react.append(list_accts)
    list_accts = []



merged = list(itertools.chain.from_iterable(final_list_react))
unique_list = []
for i in merged:
    unique_list.append(i["account"])
unique_list = list(dict.fromkeys(unique_list))

final_list_reactions = []

links = []
for i in unique_list:
    for x in merged:
        if(i == x["account"]):
            d = {}
            d["account"] = i
            links.append(x["link"])
    d["links_to_react"] = links
    final_list_reactions.append(d)
    links = []




final_list_share = []

for i in links_sheet:
    accounts_that_will_share = random.sample(number_of_accounts,int(i["Shares"]))
    for x in accounts_that_will_share:
        d = {}
        d["account"] = x
        d["link"] = i["Links"]
        list_accts.append(d)
    final_list_share.append(list_accts)
    list_accts = []



merged = list(itertools.chain.from_iterable(final_list_share))


unique_list = []
for i in merged:
    unique_list.append(i["account"])
unique_list_1 = list(dict.fromkeys(unique_list))

final_list_shares = []

links = []
for i in unique_list_1:
    for x in merged:
        if(i == x["account"]):
            d = {}
            d["account"] = i
            links.append(x["link"])
    d["links_to_share"] = links
    final_list_shares.append(d)
    links = []



# print final_list_reactions
# print final_list_shares

final_list = []

for i in range(1,int(settings_sheets[0]["Accounts"]+1)):
    d = {}
    for x in final_list_reactions:
        if(i == x["account"]):
            d["links_to_react"] = x["links_to_react"]
    for x in final_list_shares:
        if(i == x["account"]):
            d["links_to_share"] = x["links_to_share"]
    d["accounts"] = i
    final_list.append(d)

start = settings_sheets[0]["Start"]
cps = settings_sheets[0]["CP"]

print("CP" + '\t' + 'Account #' + '\t' + "Reactions" + '\t' + "Links to React" + '\t'  + "Shares" + '\t' + "Links to Share")
for i in final_list:
    i["reactions"] = 'yes'
    i["share"] = 'yes'
    try:
        i["links_to_react"]
    except:
        i["links_to_react"] = ''
        i["reactions"] = 'no'
    try:
        i["links_to_share"]
    except:
        i["links_to_share"] = ''
        i["share"] = 'no'

    if start<=cps:
        print("CP" + str(int(start)) + '\t' + str(i["accounts"]) + '\t' + i["reactions"] +'\t'+  str(i["links_to_react"])[1:-1].replace("u'", "").replace("'", "")  + '\t' + i["share"] + '\t'+ str(i["links_to_share"])[1:-1].replace("u'", "").replace("'", ""))
    else:
        start = settings_sheets[0]["Start"]
        print("CP" + str(int(start)) + '\t' + str(i["accounts"]) + '\t' + i["reactions"] +'\t'+  str(i["links_to_react"])[1:-1].replace("u'", "").replace("'", "")  + '\t' + i["share"] + '\t'+ str(i["links_to_share"])[1:-1].replace("u'", "").replace("'", ""))
    start = start + 1
